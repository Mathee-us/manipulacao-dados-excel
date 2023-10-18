from openpyxl.chart import LineChart, reference, Reference, Series
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill,fills


class Leitor:
    def __init__(self, caminho_arq: str = ""):
        self.caminho_arq = caminho_arq
        self.dados = []

    def ler_arquivo(self, acao: str = ""):
        with open(f'{self.caminho_arq}{acao}.txt', 'r') as arquivo_cotacao:
            linhas = arquivo_cotacao.readlines()
            self.dados= [linha.replace("\n", "").split(";") for linha in linhas]



class PropriedadesGrafico:
    def __init__(self, grossura: int, cor: str):
        self.grossura = grossura
        self.cor = cor


class GerenciadorPlanilha:
    def __init__(self):
        self.wb = Workbook()
        self.planilha_ativa = None

    def criar_planilha(self, nome_planilha: str = ""):
        nova_planilha = self.wb.create_sheet(nome_planilha)
        self.wb.active = nova_planilha
        self.planilha_ativa = nova_planilha

        return nova_planilha

    def adiciona_linha(self, linha=None):
        self.planilha_ativa.append(linha)

    def atualiza_cell(self, cell: str = "", valor:str = ""):
        self.planilha_ativa[cell] = valor

    def mescla_cells(self, cell_ini: str = "", cell_fim: str = ""):
        self.planilha_ativa.merge_cells(f'{cell_ini}:{cell_fim}')

    def aplica_estilo(self, celula: str, estilos: list):
        for estilo in estilos:
            atributo, valor = estilo
            setattr(self.planilha_ativa[celula], atributo, valor)

    def adiciona_grafico(self, comprimento: float, altura: float, titulo: str,
                         titu_x: str, titu_y: str, referencia_x: reference,
                         referencia_y: reference, propriedades_graf: list, cell: str):
        grafico = LineChart()
        grafico.width = comprimento
        grafico.height = altura
        grafico.title = titulo
        grafico.x_axis.title = titu_x
        grafico.y_axis.title = titu_y

        grafico.add_data(referencia_x)
        grafico.set_categories(referencia_y)


        for serie, propriedades in zip(grafico.series, propriedades_graf):
            serie.graphicalProperties.line.solidFill = propriedades.cor
            serie.graphicalProperties.line.width = propriedades.grossura

        self.planilha_ativa.add_chart(grafico, cell)

        return grafico

    def salva_planilha(self, nome_arquivo: str = ""):
        self.wb.save(f'./saida/Planilha{nome_arquivo}.xlsx')


